from openpyxl import load_workbook
from os import listdir
from os.path import isfile, join
import urllib.request
import urllib.parse
import re

# This class is defined to represent each individual track
class Track(object):
    name = ""
    votes = 0
    appearences = 0
    total = 0

    # The class "constructor" - It's actually an initializer
    def __init__(self, track):
        self.name = track
        self.points = 0
        self.appearences = 0
        self.total = 0

# Here's my global list of tracks to hold everything we find
tracks = []

# Set this path to wherever you are holding your spreadsheets
mypath = '/Users/darrylwright/projects/Spectrum Culture/'

# Format of spreadsheets is expected to be 2 columns, "Track" and "Rank" where rank is a number between 1-50.
# Anything other than this won't work.

def youtubeLookup( url ):
    query_string = urllib.parse.urlencode({"search_query": url})

    html_content = urllib.request.urlopen("http://www.youtube.com/results?" + query_string)

    search_results = re.findall(r'href=\"\/watch\?v=(.{11})', html_content.read().decode())

    if search_results[0]:
        return str.format( "http://www.youtube.com/watch?v={0}", search_results[0])

    return "NO VIDEO FOUND"

# Function to score a particular track based on the rank passed in.
def scoreTrack( track, rank ):
    if rank > 50 or rank < 1:
        return

    track.appearences += 1
    points = (rank - 51) * -1
    track.votes += points
    #print("Vote for track " + track.name + " makes it " + str(track.votes))


# Function to vote for a track with a rank
def voteTrack(trackName, rank):
    for track in tracks:
        if track.name == trackName:
          #  print("Vote for track " + trackName)
            scoreTrack( track, rank )
            return # We return early because it was found already and we ranked it.

    # If not found, we create a new track, add it to the list and rank it.
    newtrack = Track( trackName.strip() )
    print( str.format( "Vote {0} for track {1}", rank, trackName ) )
    tracks.append( newtrack )
    scoreTrack( newtrack, rank )


# Process the workbook (spreadsheet) we've read in.
def processWorkBook( wb ):
    ws = wb.active
    wb.guess_types = True

    # Look for duplicate votes
    votes = []
    for row in range(2, 261):
        cell = 'A' + str(row)
        if ws[cell].value != None and ws['B' + str(row)].value != None:
            track = ws[cell].value
            vote = ws['B' + str(row)].value
            if vote > 50:
                continue
            if vote in votes:
                print( "Error: 2 votes for " + str(vote) + " detected.")
            else:
                votes.append( vote )

    # We go through every row and check the validity of the cells for column A + B before passing them on to voteTrack
    for row in range( 2, 261 ):
        cell = 'A' + str( row )
        if ws[cell].value != None and ws['B'+str(row)].value != None:
            track = ws[cell].value
            vote = ws['B'+str(row)].value
            if len( track ) > 3:
                voteTrack(track, vote)

# ENTER HERE.
def main():
    # Get the spreadsheets
    onlyfiles = [f for f in listdir(mypath) if isfile(join(mypath, f))]

    # Go through the sheets only (skipping Microsoft's awful temp files) and process each workbook
    for xlsfile in onlyfiles:
        if xlsfile.find( ".xlsx" ) != -1 and xlsfile[0] != '~':
            print( "Processing " + xlsfile + "...")
            wb = load_workbook(filename = xlsfile)

            processWorkBook( wb )

    # Multiply the track's votes by the number of times it appeared on a list.
    for track in tracks:
        track.total = track.votes * track.appearences

    # A few handy sorts of the list
    newlist = sorted(tracks, key=lambda x: x.total, reverse=True)
    alpha = sorted(tracks, key=lambda x: x.name, reverse=False)

    print( "-----------------------------------------------------------------" )
    print( "ALPHABETICALLY")
    print("-----------------------------------------------------------------")
    for track in alpha:
        print(track.name + " has a score of " + str(track.total))

    print("\n")
    print("-----------------------------------------------------------------")
    print("BY SCORE")
    print("-----------------------------------------------------------------")
    for track in newlist:
        print( track.name + " has a score of " + str( track.total ) )

    print("\n")
    print("-----------------------------------------------------------------")
    print("TOP 100")
    print("-----------------------------------------------------------------")
    del newlist[100:]
    item = 1
    for track in newlist:
        query = track.name.lower()
        query = re.sub( r'\W+', ' ', query )

        # query = query.replace( '"', "" )
        # query = query.replace('“', "" )
        # query = query.replace(":", " ")
        # query = query.replace("-", " ")
        query += " category:Music"
        #print( query )

        link = youtubeLookup(query)
        if not link:
            link.format( "Nothing found for query \"{0}\"", query )
            print( link )

        print( str.format( "{0}: {1}   SCORE: {2} LINK: {3}", item, track.name, track.total, link) )
        item += 1




if __name__ == "__main__":
    main()

